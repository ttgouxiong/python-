import time

import openpyxl
import random
from openpyxl.styles import PatternFill

"""
将数据写入表格
"""

# 设置文件 mingc
name = "tt.xlsx"

# 打开文件
# wb = load_workbook(addr)

# 创建文件(这里一定不要写名字tmd)
wb = openpyxl.Workbook()
print(1)
# 创建一张新表
# ws= wb.active   #打开一个sheet
# ws.title='new sheet' #改默认sheet的名字为‘new sheet’
ws = wb.create_sheet('tt_test')
ws1 = wb.create_sheet('tt_test1')
ws2 = wb.create_sheet('tt_test2')
# 第一行输入
ws.append(['编号', '姓名', '年龄'])
ID = '编号'
NAME = '姓名'
AGE = '年龄'
# 输入内容(500行数据)

ws.append([1,2,3])


t = []
for i in range(5):
    ID = str(i)
    NAME = str('wenjing' + str(i))
    AGE = random.randint(18, 26)
    t += [ID, NAME, AGE]
    ws.append(t)

# 改变 sheet 标签按钮颜色
# ws.sheet_properties.tabColor = "1072BA"
green_fill = PatternFill(start_color="AACF91", end_color="AACF91", fill_type="solid")

ws['A8'].fill = green_fill
#####################################
d=ws.cell(5,5).fill = green_fill
#####################################


wb.save(name)






