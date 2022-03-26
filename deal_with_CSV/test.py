import random
from openpyxl import Workbook
wb=Workbook() #新创建Excel
ws=wb.active   #打开一个sheet
ws.title='new sheet' #改默认sheet的名字为‘new sheet’
ws1=wb.create_sheet('first')  #新创建一个sheet，并重新命名为first
ws2 = wb.create_sheet("Mysheet", 1) #默认依次在后新建，可以更改默认值
ws.sheet_properties.tabColor = "1072BA"  #设置sheet标题的背景色
source=wb.active
target=wb.copy_worksheet(source) #创建source的副本

#表格数据
#指定sheet的单元格赋值
c=ws['A5']
ws['A5']='Python'
wb.save('eee.xlsx')